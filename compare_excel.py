import openpyxl
import pytest


def read_row_column( expected, actual):
    wb1 = openpyxl.load_workbook(expected, data_only=True)
    sheet1 = wb1.active
    wb2 = openpyxl.load_workbook(actual, data_only=True)
    sheet2 = wb2.active
    expected_data = list(sheet1.iter_rows(values_only=True))
    actual_data = list(sheet2.iter_rows(values_only=True))
    expected_headers = [header.strip() if header else "" for header in expected_data[0]]
    actual_headers = [header.strip() if header else "" for header in actual_data[0]]
    expected_dict = {row[0]: {expected_headers[i]: row[i] for i in range(1, len(row))} for row in expected_data[1:]}
    actual_dict = {row[0]: {actual_headers[i]: row[i] for i in range(1, len(row))} for row in actual_data[1:]}

    for id_key, expected_row in expected_dict.items():
        if id_key not in actual_dict:
            raise AssertionError(f"ID {id_key} not found in actual data")

        actual_row = actual_dict[id_key]
        expected_columns = expected_row.get('Expected Column', "").split(',')

        for col in expected_columns:
            col = col.strip()
            expected_value = expected_row.get(col, None)
            actual_value = actual_row.get(col, None)

            assert expected_value == actual_value, (
                f"Mismatch for ID {id_key} in column '{col}': "
                f"Expected: {expected_value}, Actual: {actual_value}"
            )


@pytest.mark.parametrize("expected_sheet,actual_sheet", [
    ("5.xlsx","6.xlsx")])
def test_data(expected_sheet, actual_sheet):
    read_row_column(expected_sheet, actual_sheet)

